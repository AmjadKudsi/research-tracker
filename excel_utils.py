import openpyxl
from openpyxl import load_workbook
import pandas as pd

EXCEL_PATH = "Research.xlsx"

def load_sheet():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)
    wb.close()  
    return df


def append_row_to_excel(entry: dict):
    """
    Append a new paper row to the bottom of the sheet.
    Columns expected in Excel:
    A Name
    B Category
    C Date
    D Notes
    E Link
    F Security (Yes/No)
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1, value=entry["Name"])
    ws.cell(row=next_row, column=2, value=entry["Category"])
    ws.cell(row=next_row, column=3, value=entry["Date"])
    ws.cell(row=next_row, column=4, value=entry["Notes"])
    ws.cell(row=next_row, column=5, value=entry["Link"])
    ws.cell(row=next_row, column=6, value=entry["Security"])  # <-- new

    wb.save(EXCEL_PATH)
    wb.close()


def get_metrics(df: pd.DataFrame) -> dict:
    """
    Values for the KPI cards.
    """

    # drop fully empty rows from df first
    df_clean = df.dropna(how="all")

    total_papers = len(df_clean)

    # ---- phase coverage metrics ----
    if "Category" in df_clean.columns:
        phase_counts = (
            df_clean["Category"]
            .value_counts(dropna=True)
            .to_dict()
        )
    else:
        phase_counts = {}

    all_phases = [
        "Phase 1",
        "Phase 2",
        "Phase 3",
        "Phase 4",
        "Phase 5",
        "Phase 6",
        "Phase 7",
        "Phase 8",
        "Phase 9",
        "Phase 10",
    ]

    covered = [p for p in all_phases if phase_counts.get(p, 0) > 0]
    coverage_text = f"{len(covered)} / {len(all_phases)} phases covered"

    # ---- security metrics ----
    if "Security" in df_clean.columns:
        security_mask = df_clean["Security"].astype(str).str.strip().str.lower() == "yes"
        security_papers = int(security_mask.sum())
    else:
        security_papers = 0

    if total_papers > 0:
        pct = round((security_papers / total_papers) * 100)
        security_ratio_pct = f"{pct}% of total"
    else:
        security_ratio_pct = "0% of total"

    # ---- bottom phase cards info ----
    phase_gap_info = []
    for phase in all_phases:
        count = phase_counts.get(phase, 0)
        phase_gap_info.append({
            "phase": phase,
            "count": count,
        })

    return {
        "total_papers": total_papers,
        "phase_coverage_text": coverage_text,
        "phase_gap_info": phase_gap_info,
        "security_papers": security_papers,
        "security_ratio_pct": security_ratio_pct,
    }